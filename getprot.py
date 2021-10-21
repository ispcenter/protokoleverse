import shutil, os, xlrd, openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment

print('GetProt started')
# class Protokol:
# 	def __init__(self):
# 		self.serialNumber = 
# 		self.model = 
# 		self.

# class Priemka:

# 	def __init__(self, date, protCount):
# 		self.date = date
# 		self.protCount = protCount

# 		if protCount <= 1:
# 			prot_1 = Protokol()


def protoFinder(serialNumber, path, result):

	for root, dirs, files in os.walk(path):
		for file in files:
			if serialNumber in file:
				pathf = os.path.join(root, file)
				result.append(pathf)


def xls_to_xlsx(file):
	# 1 Read xls
	wb1 = xlrd.open_workbook(file)
	ws1 = wb1.sheet_by_index(0)
	data = []
	
	for row in ws1:
		rowList = []
		for cell in row:
			rowList.append(cell.value)
		data.append(rowList)

	# 2 Create openpyxl wb
	wb2 = openpyxl.Workbook()
	ws = wb2.active

	for row in data:

		if (len(row[0]) != 0 or len(row[4]) != 0):
			if any([
				row[0].startswith('Испытатель'),
				row[0].startswith('Начальник'),
				row[0].startswith('Главный'),
				row[0].startswith('ЗАКЛЮЧЕНИЕ'),
				row[0].startswith('Прим'),
				row[0].startswith('Представитель')]):
				ws.append([])
			ws.append(row)


	# 3 Set styles
	for row in ws['A1:A4']:
		for cell in row:
			cell.style = Heading

	for row in ws['A5:F13']:
		for cell in row:
			cell.style = TabledC

	for obl in [ws['B16:F16'], ws['B19:F19'], ws['B26:F26'], ws['B29:F29']]:
		for row in obl:
			for cell in row:
				cell.style = underscore

	for obl in [ws['C15:E15'], ws['C18:E18'], ws['C25:E25'], ws['C28:E28']]:
		for row in obl:
			for cell in row:
				cell.style = Heading

	ws['A21'].style = Heading

	for row in ws['A32:F34']:
		for cell in row:
			cell.style = underscore

	ws.merge_cells('A1:F1')
	ws.merge_cells('A2:F2')
	ws.merge_cells('A3:F3')
	ws.merge_cells('A4:F4')

	ws.merge_cells('A5:A6')
	ws.merge_cells('B5:B6')
	ws.merge_cells('C5:C6')
	ws.merge_cells('D5:E5')
	ws.merge_cells('F5:F6')

	ws.merge_cells('A12:C12')
	ws.merge_cells('A13:C13')

	ws.merge_cells('D12:E12')
	ws.merge_cells('D13:E13')

	ws.merge_cells('A21:F21') # заключение
	ws.merge_cells('A22:F22')
	ws.merge_cells('A23:F23')

	ws.merge_cells('C15:D15')
	ws.merge_cells('E15:F15')
	ws.merge_cells('C16:D16')
	ws.merge_cells('E16:F16')

	ws.merge_cells('C18:D18')
	ws.merge_cells('E18:F18')
	ws.merge_cells('C19:D19')
	ws.merge_cells('E19:F19')

	ws.merge_cells('C25:D25')
	ws.merge_cells('E25:F25')
	ws.merge_cells('C26:D26')
	ws.merge_cells('E26:F26')

	ws.merge_cells('C28:D28')
	ws.merge_cells('E28:F28')
	ws.merge_cells('C29:D29')
	ws.merge_cells('E29:F29')

	ws['A25'].alignment = Alignment(wrap_text=True)
	k = 1.03
	ws.column_dimensions['A'].width = 26.29 * k
	ws.column_dimensions['B'].width = 8.29 * k
	ws.column_dimensions['C'].width = 11.29 * k
	ws.column_dimensions['D'].width = 10.29 * k
	ws.column_dimensions['E'].width = 10.29 * k
	ws.column_dimensions['F'].width = 10.29 * k

	ws.row_dimensions[1].height = 15.0
	ws.row_dimensions[2].height = 25.5
	ws.row_dimensions[3].height = 25.5
	ws.row_dimensions[5].height = 39
	ws.row_dimensions[6].height = 25.5

	ws.row_dimensions[7].height = 25.5
	ws.row_dimensions[8].height = 25.5
	ws.row_dimensions[9].height = 25.5
	ws.row_dimensions[10].height = 25.5
	ws.row_dimensions[11].height = 25.5
	ws.row_dimensions[12].height = 25.5

	ws.row_dimensions[31].height = 25.5
	ws.row_dimensions[32].height = 25.5
	ws.row_dimensions[33].height = 25.5

	# 4 Check data
	ws['A1'].value = 'Протокол № 2299/21'

	# 5 Rename
	fileName = file[:-30] # cut date and other rabbish
	prType = fileName[:2]
	model = fileName[3:-7]
	serialNumber = fileName[-6:]
	protNumber = ws['A1'].value.split('/')[0].split(' ')[-1]
	fileName = f'{protNumber}_{prType}_{model}_{serialNumber}.xlsx'

	# 6 Save xlsx file
	wb2.save(fileName)

bdr_thin = Side(style='thin', color="000000")

# стиль "TabledC"
global TabledC
TabledC = NamedStyle(name="TabledC")
TabledC.border = Border(left=bdr_thin, top=bdr_thin, right=bdr_thin, bottom=bdr_thin)
TabledC.alignment = Alignment(vertical = 'center', horizontal = 'center', wrap_text=True)
TabledC.font = Font(name='Arial', size=10)

# стиль Заголовок
global Heading
Heading = NamedStyle(name="Heading")
Heading.alignment = Alignment(vertical = 'bottom', horizontal = 'center')
Heading.font = Font(name='Arial', size=10, bold=True)

# создадим стиль "underscore"
global underscore
underscore = NamedStyle(name="underscore")
underscore.border = Border(top=bdr_thin)
underscore.alignment = Alignment(vertical = 'top', horizontal = 'center')
underscore.font = Font(name='Arial', size=6)

# 1 Read the task

with open('numbers.csv', 'r') as file:
	serialNumbers = file.read().split(';')

path = r'\\Saukip\arm_3\Аттестация\2021\октября'
result = []

for serialNumber in serialNumbers:
	protoFinder(serialNumber, path, result)

# 2 Get data from server

# 2.1 Create workFolder
workFolder = os.getcwd() + '\\protFolder'
try:
	os.makedirs(workFolder)
except FileExistsError:
	pass

os.chdir(workFolder)

# 2.2 Copy nesessary files
for file in result:
	shutil.copy2(file, file.split('\\')[-1])
	print(file)


# 2.3 Convert xls into xlsx
for file in os.listdir(os.getcwd()):
	xls_to_xlsx(file)

# Check protokol data

# Print protokols by printer
# for i in list
# os.startfile(i, 'print')

print('GetProt finished')