from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import openpyxl

# 1 add custom method to openpyxl...Cell class
def changeCellType(self):
  if self.value < 10:
    self.fill = badBG
    self.font = badFont
  if self.value >= 10:
    self.fill = goodBG
    self.font = goodFont

openpyxl.cell.cell.Cell.changeCellType = changeCellType

# 2 add custom cell styles
badBG = PatternFill(start_color="EB9C9C", end_color="EB9C9C", fill_type = "solid")
goodBG = PatternFill(start_color="9CEBB5", end_color="9CEBB5", fill_type = "solid")
badFont = Font(name='Calibri', size=12, color="770808")
goodFont = Font(name='Tahoma', size=12, color="06541F")

# 3 create workbook object
wb = Workbook()
ws = wb.active
ws['A1'].value = 100
ws['A2'].value = 99

# 4 change cell type
for row in ws['A1:A2']:
  for cell in row:
    cell.changeCellType()

wb.save('testcol.xlsx')
